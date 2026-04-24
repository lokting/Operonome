import os
import re
import openpyxl as xl

gff_dict = {}
gff_list_big = []
seqname_list = []

input_gff_dir = r"input/genome_gff/"  
input_gff_file_paths  = [f for f in os.listdir(input_gff_dir) if f.endswith(".gff")]
input_gff_file_path = os.path.join(input_gff_dir, input_gff_file_paths[0])
output_excel_path = r'tmp_output/gff.xlsx'

with open(input_gff_file_path) as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split('\t')
        if len(parts) < 9:  
            print(f"Skipping invalid line: {line}")
            continue
        if parts[2] != "CDS":
            continue

        try:
            seqname = parts[0]
            Start_coordinate = int(parts[3])
            End_coordinate = int(parts[4])
            Strand = parts[6]
            attributes = parts[8]
            
            locus_match = re.search(r'locus_tag=([^;]+)', attributes)
            if not locus_match:
                print(f"Missing locus_tag in line: {line}")
                continue
            LocusTag = locus_match.group(1)

            Gene_length = abs(End_coordinate - Start_coordinate) + 1

            if Strand == "-":
                Start_coordinate, End_coordinate = End_coordinate, Start_coordinate

            if seqname not in seqname_list:
                gff_list_big = []
                seqname_list.append(seqname)
            
            gff_list_samll = [
                seqname,
                Start_coordinate,
                End_coordinate,
                Strand,
                LocusTag,
                Gene_length
            ]
            gff_list_big.append(gff_list_samll)
            gff_dict[seqname] = gff_list_big

        except (ValueError, IndexError) as e:
            print(f"Error processing line: {line}\nError: {e}")
            continue

workbook_excel = xl.Workbook()
for seqname, entries in gff_dict.items():
    negative_strand = sorted(
        [e for e in entries if e[3] == "-"],
        key=lambda x: x[1],
        reverse=True  
    )
    positive_strand = [e for e in entries if e[3] == "+"]

    for strand_type, data in [("+", positive_strand), ("-", negative_strand)]:
        sheet_name = f"{seqname}_{strand_type}"
        workbook_excel.create_sheet(sheet_name)
        ws = workbook_excel[sheet_name]
        ws.append(["seqname", "Start", "End", "Strand", "LocusTag", "Length", "Intergenegap"])
        
        prev_end = None
        for row in data:
            if prev_end is not None:
                gap = row[1] - prev_end - 1
                row.append(gap if gap > 0 else 0)
            else:
                row.append(0)
            prev_end = row[2]
            ws.append(row)

if "Sheet" in workbook_excel.sheetnames:
    del workbook_excel["Sheet"]
           
workbook_excel.save(output_excel_path)
print(f"文件已成功保存至：{output_excel_path}")
