#警告 无法完成闭环 可能需要手动处理
#在结果中，搜索每张表最后一个localtag，只有出现 最后一个localtag-第一个localtag  第一个localtag-第二个localtag （或者紧接着的连接）
#后者拼接上前者  手动更新  同时更新operon顺序
#所以只要搜索每张表第一个localtag，只要不同时出现 最后一个localtag-第一个localtag  第一个localtag-第二个localtag 就说明首尾不连接，可不干预先

#只求chromosome的operon

#多重条件组合技  计算operon与基因的绑定关系
import openpyxl as xl


input_pearson_num_txt = r"tmp_output/gene_PCC.txt"
input_excel_path = r"tmp_output/gff.xlsx" 
output_operon_gene = r'output/operon-gene.xlsx'
chr_list_file = r'input/chr_list.txt'

def read_list_from_file(file_path):
    with open(file_path, 'r') as file:
        return [line.strip() for line in file]
    
#染色体sheetID
chr_list = read_list_from_file(chr_list_file)

#构造{localtag1-localtag2:pearson_num}
pearson_num_dict = {}
with open(input_pearson_num_txt) as pearson_num:
    for i in pearson_num:
        line = i.split("\t")
        localtag_count = line[0]
        pearson_num_count = line[1]
        pearson_num_dict[localtag_count] = pearson_num_count

#遍历excel表
inter_localtag_list_small = []
inter_localtag_dict = {}
operon_num = 1
workbook = xl.load_workbook(input_excel_path)
# for sheet_name in workbook.sheetnames:
for sheet_name in chr_list:
    worksheet = workbook[sheet_name]
    num = 1
    finial = worksheet.max_row #表格数量
    while num<finial:
        operon = "openon_" + str(operon_num)
        num = num + 1

        #取值
        if num == finial:
            cell_last_localtag = worksheet["E"+str(num)]
            last_localtag = cell_last_localtag.value
            cell_next_localtag = worksheet["E"+str(2)]
            next_localtag = cell_next_localtag.value
            inter_localtag = last_localtag + "-" + next_localtag
            pearson_num_inter = float(pearson_num_dict[inter_localtag]) #获得PCC(float)
            cell_intergap = worksheet["G"+str(num)]
            intergap = int(cell_intergap.value) #获得基因间距（int）
        else:
            cell_last_localtag = worksheet["E"+str(num)]
            last_localtag = cell_last_localtag.value
            cell_next_localtag = worksheet["E"+str(num + 1)]
            next_localtag = cell_next_localtag.value
            inter_localtag = last_localtag + "-" + next_localtag
            pearson_num_inter = float(pearson_num_dict[inter_localtag]) #获得PCC(float)
            cell_intergap = worksheet["G"+str(num)]
            intergap = int(cell_intergap.value) #获得基因间距（int）


        #判断
        #PCC=nan 说明其中一个是假想基因 认为二者没有联系
        if pearson_num_inter == "nan": 
            #锁定operon组合 去重 保留顺序
            localtag_new_list = []
            if inter_localtag_list_small != []:
                for localtag_new in inter_localtag_list_small:
                    if localtag_new not in localtag_new_list:
                        localtag_new_list.append(localtag_new)
                inter_localtag_dict[operon] = localtag_new_list
            localtag_new_list = []
            inter_localtag_list_small = []
            operon_num = operon_num + 1
        #多条件组合 只要满足其中一个条件，就认为两个基因有关联
        elif (intergap <= 10) or (intergap <= 50 and pearson_num_inter > 0.5)or (intergap <= 100 and pearson_num_inter > 0.6)or (intergap <= 200 and pearson_num_inter > 0.7)or (intergap <= 500 and pearson_num_inter > 0.8):
            inter_localtag_list_small.append(last_localtag)
            inter_localtag_list_small.append(next_localtag)
            #最后一个
            if num == finial:
                #锁定operon组合 去重 保留顺序
                localtag_new_list = []
                if inter_localtag_list_small != []:
                    for localtag_new in inter_localtag_list_small:
                        if localtag_new not in localtag_new_list:
                            localtag_new_list.append(localtag_new)
                    inter_localtag_dict[operon] = localtag_new_list
                localtag_new_list = []
                inter_localtag_list_small = []
                operon_num = operon_num + 1
        #不满足条件的 认为二者没有联系
        else:
            #锁定operon组合 去重 保留顺序
            localtag_new_list = []
            if inter_localtag_list_small != []:
                for localtag_new in inter_localtag_list_small:
                    if localtag_new not in localtag_new_list:
                        localtag_new_list.append(localtag_new)
                inter_localtag_dict[operon] = localtag_new_list
            localtag_new_list = []
            inter_localtag_list_small = []
            operon_num = operon_num + 1


#读取上述已经分组的localtag集合
localtag_many_list = []
localtag_output_list = []
for key_1,value_1 in inter_localtag_dict.items():
    localtag_output_list.append(value_1)
    localtag_many_list.extend(value_1)

#读取excel文件中所有localtag
localtag_all_list = []
workbook = xl.load_workbook(input_excel_path)
# for sheet_name in workbook.sheetnames:
for sheet_name in chr_list :
    worksheet = workbook[sheet_name]
    num = 1
    finial = worksheet.max_row #表格数量
    while num<finial:
        num = num + 1
        cell_localtag = worksheet["E"+str(num)]
        localtag = cell_localtag.value
        localtag_all_list.append(localtag)
print("localtag_all_list")
print(len(localtag_all_list))


#作差，提取出非多基因组合的基因
single_localtag_list = list( set(localtag_all_list) - set(localtag_many_list) )

#写入数据  重置operon的顺序 从1开始
localtag_output_list.extend(single_localtag_list)
print("localtag_output_list")
print(len(localtag_output_list))

str_new = ","
workbook_excel = xl.Workbook()
workbook_excel.create_sheet("Sheet")
worksheeet_excel = workbook_excel["Sheet"]
worksheeet_excel.append(["operon","localtag"])
operon_num_new = 1
for localtag_output in localtag_output_list:
    operon_name_new = "operon_" + str(operon_num_new)
    if type(localtag_output) == list:
        localtag_output_new = str_new.join(localtag_output)
    else:
        localtag_output_new = localtag_output
    worksheeet_excel.append([operon_name_new,localtag_output_new])
    operon_num_new = operon_num_new + 1
#保存结果               
workbook_excel.save(output_operon_gene)



