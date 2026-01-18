# 对4-0的 operon-gene.xlsx 进行功能注释 注释补充完整

import openpyxl as xl

#定义函数：将excel中 operon对应的 localtag串 以 列表 形式返回
def get_localtag_list(localtag_str):
    if "," in localtag_str:
        list_output = localtag_str.split(",")
    else:
        list_output = [localtag_str]
    return list_output


input_operon_gene = r'output/operon-gene.xlsx'
input_gbff_excel = r"tmp_output/gbff_function.xlsx"


#读取gbff_function文件，获取 localtag-product 字典
localtag_product = {}
workbook_2 = xl.load_workbook(input_gbff_excel)
worksheet_2 = workbook_2["Sheet1"]
num_2 = 1
finial_2 = worksheet_2.max_row #表格数量
while num_2<finial_2:
    num_2 = num_2 + 1

    # 读取值
    cell_Local_tag_2 = worksheet_2["A"+str(num_2)]
    Localtag_2 = cell_Local_tag_2.value
    cell_product_2 = worksheet_2["B"+str(num_2)]
    product_2 = cell_product_2.value

    localtag_product[Localtag_2] = product_2

#读取operon_gene文件，依次给予功能注释
workbook = xl.load_workbook(input_operon_gene)
worksheet = workbook["Sheet"]
num = 1
finial = worksheet.max_row #表格数量
worksheet["C1"] = "function" #补一个表头
while num < finial:
    num = num + 1

    # 读取值
    cell_operon = worksheet["A"+str(num)]
    operon = cell_operon.value
    cell_localtag = worksheet["B"+str(num)]
    localtag = cell_localtag.value

    #获取localtag列表
    localtag_list = get_localtag_list(localtag)

    #查找localtag功能并注释
    function_str = ""
    for localtag_check in localtag_list:
        prot_output = localtag_product.get(localtag_check,"0") #如果localtag没有对应功能注释，设为0
        function_str = function_str + prot_output + ";"
    function_str = function_str[:-1]

    #赋值
    worksheet["C"+str(num)] = function_str

workbook.save(input_operon_gene)
