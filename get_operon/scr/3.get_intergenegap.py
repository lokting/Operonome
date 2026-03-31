import os
import re
import openpyxl as xl

def read_genome_lengths(file_path):
    genome_lengths = {}
    with open(file_path, 'r') as file:
        for line in file:
            if '=' in line:
                name, length = line.strip().split('=')
                genome_lengths[name.strip()] = int(length.strip())
    return genome_lengths

def get_genome_length(name, genome_lengths):
    for key in genome_lengths:
        if key in name:
            return genome_lengths[key]
    return None

def sort_sheet_by_column(worksheet, column_name):
    col_idx = None
    for idx, cell in enumerate(worksheet[1]):
        if cell.value == column_name:
            col_idx = idx + 1
            break
    
    if col_idx:
        # Extract rows and sort
        rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        sorted_rows = sorted(rows, key=lambda row: row[col_idx - 1] or 0, reverse=True)
        
        # Write sorted rows back
        for row_idx, row in enumerate(sorted_rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

input_excel_path = r"tmp_output/gff.xlsx" 
genome_length_file = r"input/genome_length.txt"
genome_lengths = read_genome_lengths(genome_length_file)

workbook = xl.load_workbook(input_excel_path)
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    if "+" in sheet_name:
        num = 1
        finial = worksheet.max_row 
        while num<finial:
            num = num + 1

            if num == finial:
                cell_last_end_num = worksheet["C"+str(num)]
                last_end_num = int(cell_last_end_num.value)
                cell_next_start_num = worksheet["B"+str(2)]
                next_start_num = int(cell_next_start_num.value)
                genemone_length = get_genome_length(sheet_name,genome_lengths)
                print(genemone_length)
                intergapnum = genemone_length - last_end_num + next_start_num - 1

            else:
                cell_last_end_num = worksheet["C"+str(num)]
                last_end_num = int(cell_last_end_num.value)
                cell_next_start_num = worksheet["B"+str(num+1)]
                next_start_num = int(cell_next_start_num.value)
                if next_start_num > last_end_num :
                    intergapnum = next_start_num - last_end_num - 1
                if next_start_num == last_end_num :
                    intergapnum = 0
                if next_start_num < last_end_num:
                    intergapnum = next_start_num - last_end_num + 1
            
            worksheet["G"+str(num)] = intergapnum
    
    if "-" in sheet_name:
        sort_sheet_by_column(worksheet, "Start_coordinate")

        num = 1
        finial = worksheet.max_row 
        while num<finial:
            num = num + 1

            if num == finial:
                cell_last_end_num = worksheet["C"+str(num)]
                last_end_num = int(cell_last_end_num.value)
                cell_next_start_num = worksheet["B"+str(2)]
                next_start_num = int(cell_next_start_num.value)
                genemone_length = get_genome_length(sheet_name,genome_lengths)
                intergapnum = genemone_length - next_start_num + last_end_num - 1

            else:
                cell_last_end_num = worksheet["C"+str(num)]
                last_end_num = int(cell_last_end_num.value)
                cell_next_start_num = worksheet["B"+str(num+1)]
                next_start_num = int(cell_next_start_num.value)
                if last_end_num > next_start_num :
                    intergapnum = last_end_num - next_start_num - 1
                if last_end_num == next_start_num :
                    intergapnum = 0
                if last_end_num < next_start_num :
                    intergapnum = last_end_num - next_start_num + 1

            worksheet["G"+str(num)] = intergapnum
             
workbook.save(input_excel_path)
